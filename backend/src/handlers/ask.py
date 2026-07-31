import json
import os
import uuid
import base64
import io
from datetime import datetime
import urllib.request
import urllib.error
import boto3
from services.db_service import track_question_frequency

dynamodb = boto3.resource('dynamodb')
TABLE_NAME = os.environ.get('TABLE_NAME', 'StudentQuestionsTable')
table = dynamodb.Table(TABLE_NAME)

GROQ_API_KEY = os.environ.get('OPENAI_API_KEY', '')

CATEGORY_PROMPTS = {
    "Tuition & Fees": (
        "You are a Senior Student Financial Advisor. Focus on tuition schedules, "
        "bursar procedures, fee structures, payment portals, and financial aid policies."
    ),
    "Academics": (
        "You are an Academic Registrar Advisor. Focus on course credits, degree progress, "
        "add/drop deadlines, attendance policies, and academic probation guidelines."
    ),
    "Exams": (
        "You are an Examination Officer. Focus on examination rules, timetables, "
        "invigilation procedures, conflict resolution, and grading scale policies."
    ),
    "General": (
        "You are StudyPal AI, an all-round university student support assistant. "
        "Provide helpful, concise, and friendly guidance for student life."
    )
}

def extract_text_from_attachment(attachment):
    """
    Decodes Base64 file payload and extracts plain text or returns image metadata.
    Supports .txt, .pdf, .docx, .xlsx, .xls, and image files (.png, .jpg, .jpeg).
    """
    if not attachment or 'fileData' not in attachment:
        return "", None

    try:
        raw_b64 = attachment['fileData']

        # Strip Data URL header if present (e.g., "data:image/png;base64,")
        if ',' in raw_b64:
            raw_b64 = raw_b64.split(',', 1)[1]

        file_name = attachment.get('fileName', '').lower()

        # Check if attachment is an Image
        if file_name.endswith(('.png', '.jpg', '.jpeg')):
            mime_type = 'image/png' if file_name.endswith('.png') else 'image/jpeg'
            return None, {"b64": raw_b64, "mime": mime_type}

        raw_bytes = base64.b64decode(raw_b64)
        extracted_text = ""

        # 1. Plain Text (.txt)
        if file_name.endswith('.txt'):
            extracted_text = raw_bytes.decode('utf-8', errors='ignore')

        # 2. PDF Documents (.pdf)
        elif file_name.endswith('.pdf'):
            try:
                from pypdf import PdfReader
                pdf_file = io.BytesIO(raw_bytes)
                reader = PdfReader(pdf_file)
                extracted_text = "\n".join([page.extract_text() or '' for page in reader.pages])
            except ImportError:
                print("[WARN] pypdf library missing from deployment package.")
                extracted_text = "[Error: PDF parser library missing in backend]"

        # 3. Word Documents (.docx)
        elif file_name.endswith('.docx'):
            try:
                import docx
                docx_file = io.BytesIO(raw_bytes)
                doc = docx.Document(docx_file)
                extracted_text = "\n".join([p.text for p in doc.paragraphs if p.text])
            except ImportError:
                print("[WARN] python-docx library missing from deployment package.")
                extracted_text = "[Error: DOCX parser library missing in backend]"

        # 4. Excel Documents (.xlsx, .xls)
        elif file_name.endswith(('.xlsx', '.xls')):
            try:
                import openpyxl
                excel_file = io.BytesIO(raw_bytes)
                workbook = openpyxl.load_workbook(filename=excel_file, data_only=True)
                excel_rows = []

                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    excel_rows.append(f"--- Sheet: {sheet_name} ---")
                    
                    for row in sheet.iter_rows(values_only=True):
                        if not any(row):
                            continue
                        row_str = " | ".join([str(cell) if cell is not None else "" for cell in row])
                        excel_rows.append(row_str)

                extracted_text = "\n".join(excel_rows)
            except ImportError:
                print("[WARN] openpyxl library missing from deployment package.")
                extracted_text = "[Error: Excel parser library missing in backend]"

        return extracted_text.strip()[:8000], None

    except Exception as e:
        print(f"[ERROR] Failed to extract attachment content: {str(e)}")
        return "", None

def call_groq_api(system_prompt, user_question, history=None, image_b64=None, image_mime="image/png"):
    url = "https://api.groq.com/openai/v1/chat/completions"

    # Default model for standard text and documents
    model = "llama-3.3-70b-versatile"

    # 1. Start with System Prompt
    messages = [{"role": "system", "content": system_prompt}]

    # 2. Inject previous chat history if provided (limiting to last 8 turns)
    if history and isinstance(history, list):
        for msg in history[-8:]:
            role = msg.get("role", "user")
            content = msg.get("content", "")
            if content:
                messages.append({"role": role, "content": content})

    # 3. Build user content payload
    user_content = user_question

    # If an image is present, switch model and switch content to multimodal format
    if image_b64:
        model = "qwen/qwen3.6-27b"
        user_content = [
            {"type": "text", "text": user_question or "Please analyze and describe this attached image."},
            {
                "type": "image_url",
                "image_url": {
                    "url": f"data:{image_mime};base64,{image_b64}"
                }
            }
        ]

    messages.append({"role": "user", "content": user_content})

    payload = {
        "model": model,
        "messages": messages,
        "temperature": 0.5,
        "max_tokens": 1000
    }

    headers = {
        "Authorization": f"Bearer {GROQ_API_KEY}",
        "Content-Type": "application/json",
        "User-Agent": "StudyPal-AI/1.0"
    }

    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode('utf-8'),
        headers=headers,
        method='POST'
    )

    try:
        with urllib.request.urlopen(req) as response:
            res_body = json.loads(response.read().decode('utf-8'))
            return res_body['choices'][0]['message']['content']
    except urllib.error.HTTPError as e:
        error_res = e.read().decode('utf-8')
        print(f"Groq API HTTP Error {e.code}: {error_res}")
        raise Exception(f"Groq API returned status {e.code}: {error_res}")

def handler(event, context):
    headers = {
        "Content-Type": "application/json",
        "Access-Control-Allow-Origin": "*",
        "Access-Control-Allow-Headers": "Content-Type,Authorization,X-Amz-Date,X-Api-Key,X-Amz-Security-Token",
        "Access-Control-Allow-Methods": "OPTIONS,POST"
    }

    if event.get('httpMethod') == 'OPTIONS':
        return {"statusCode": 200, "headers": headers, "body": json.dumps({"message": "OK"})}

    try:
        # Extract Cognito User ID if available
        user_id = None
        try:
            user_id = event['requestContext']['authorizer']['claims']['sub']
        except (KeyError, TypeError):
            pass

        body = json.loads(event.get('body', '{}'))
        question = body.get('question', '').strip()
        category = body.get('category', 'General')
        attachment = body.get('attachment', None)
        history = body.get('history', [])  # Extract chat history array from request
        chat_id = body.get('chatId')
        chat_title = body.get('chatTitle')

        # Require at least a written question OR an attached file
        if not question and not attachment:
            return {
                "statusCode": 400,
                "headers": headers,
                "body": json.dumps({"error": "Either a question or an attached document is required."})
            }

        # Step 1: Extract document text or image payload if attachment exists
        extracted_doc_text, image_payload = extract_text_from_attachment(attachment)

        # Step 2: Build final prompt combining user text and extracted file content
        file_name = attachment.get('fileName', 'Attachment') if attachment else None

        final_prompt = question
        if extracted_doc_text:
            final_prompt += f"\n\n--- Attached Document Context ({file_name}) ---\n{extracted_doc_text}\n--- End of Document Context ---"

        system_instruction = CATEGORY_PROMPTS.get(category, CATEGORY_PROMPTS['General'])

        # Step 3: Invoke Groq API with history
        if GROQ_API_KEY and GROQ_API_KEY != "PLACEHOLDER_KEY":
            if image_payload:
                ai_answer = call_groq_api(
                    system_instruction, 
                    final_prompt, 
                    history=history,
                    image_b64=image_payload["b64"], 
                    image_mime=image_payload["mime"]
                )
            else:
                ai_answer = call_groq_api(system_instruction, final_prompt, history=history)
        else:
            ai_answer = f"[{category}] Groq API Key is not configured in environment variables."

        # Step 4: Persist Q&A Record into DynamoDB
        item_id = str(uuid.uuid4())
        created_at = datetime.utcnow().isoformat()

        item = {
            "id": item_id,
            "question": question if question else f"[File Uploaded: {file_name}]",
            "category": category,
            "answer": ai_answer,
            "createdAt": created_at,
            "hasAttachment": bool(attachment)
        }

        if file_name:
            item["fileName"] = file_name

        if user_id:
            item["userId"] = user_id

        if chat_id:
            item["chatId"] = chat_id
        if chat_title:
            item["chatTitle"] = chat_title

        # Track question frequency across platform
        if question:
            track_question_frequency(question, category)

        table.put_item(Item=item)

        return {
            "statusCode": 200,
            "headers": headers,
            "body": json.dumps({
                "data": item
            })
        }

    except Exception as e:
        print(f"[ERROR] Groq Lambda failure: {str(e)}")
        return {
            "statusCode": 500,
            "headers": headers,
            "body": json.dumps({"error": str(e)})
        }
